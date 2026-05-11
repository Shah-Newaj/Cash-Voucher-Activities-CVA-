from openpyxl import load_workbook


class ExcelUtils:

    @staticmethod
    def get_data(file_name, sheet_name):

        workbook = load_workbook(file_name)
        sheet = workbook[sheet_name]

        data = []

        for row in sheet.iter_rows(min_row=2, values_only=True):

            # Skip fully empty rows
            if all(cell is None for cell in row):
                continue

            data.append(row)

        return data